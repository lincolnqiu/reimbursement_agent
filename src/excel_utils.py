'''
excel_utils.py
==============
该文件包含所有与 Excel 文件操作相关的辅助函数。
主要功能包括：
1. 创建新的 Excel 工作簿 (`_create_workbook`)，并设置表头。
2. 将提取的发票数据写入指定的 Excel 文件 (`_write_to_excel`)，如果文件不存在则先创建它。
   此函数还负责在金额列的末尾添加求和公式。

更新条件：
- 当 Excel 报表的格式（如表头、列顺序、求和逻辑）需要修改时，更新此文件中的函数。
- 如果更换了处理 Excel 的库 (当前为 `openpyxl`)，需要重写此文件中的相关逻辑。
'''

import os
from typing import List, Dict

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# 从 config.py 导入 EXCEL_PATH，尽管在这里直接使用它可能不是最佳实践，
# 但为了保持与原代码相似的逻辑，暂时这样做。
# 更优的做法是在调用时传递路径。
from config import EXCEL_PATH

def create_workbook(path: str = EXCEL_PATH):
    """创建一个新的 Excel 工作簿并设置表头。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoices"
    # 更新列名以匹配主脚本中的期望
    ws.append(["发票类型", "金额", "类别", "发票号码", "新文件名"]) # 原为 "原文件名"
    wb.save(path)
    print(f"📄 新的 Excel 文件已创建: {path}")

def write_to_excel(data: List[Dict[str, str]], xl_path: str = EXCEL_PATH):
    """将数据写入 Excel 文件。如果文件不存在，则先创建它。"""
    if not os.path.exists(xl_path):
        create_workbook(xl_path)
    
    wb = load_workbook(xl_path)
    ws = wb.active

    start_row = ws.max_row + 1
    for item in data:
        ws.append([
            item.get("invoice_type"),
            float(item.get("amount")) if item.get("amount") and item.get("amount").replace('.','',1).isdigit() else None, # 增加校验，确保金额可转换为float
            item.get("category"),
            item.get("invoice_number"),
            item.get("file_name") # 对应 `main` 函数中的 `new_pdf_name`
        ])

    # 更新求和公式（金额列为 B，即第二列）
    amount_col_letter = get_column_letter(2) # B列
    end_row = ws.max_row
    
    # 确保有数据行才添加总计 (至少有一行数据 + 表头)
    if end_row >= 2: 
        formula_cell = f"{amount_col_letter}{end_row + 1}"
        # 求和范围从第二行数据开始 (ws.cell(row=2, column=2)) 到最后一行数据
        ws[formula_cell] = f"=SUM({amount_col_letter}2:{amount_col_letter}{end_row})"
        ws[f"A{end_row + 1}"] = "总计"
    else:
        print("ℹ️ Excel中没有数据行，未添加总计公式。")

    wb.save(xl_path)
    print(f"📊 数据已写入 {xl_path}") 